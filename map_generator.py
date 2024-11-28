"""
please install openpyxl first by enter 'pip install openpyxl' in cmd
使用方法: 在map_template.xlsx中填1的格子会生成一格墙
黑色边框是地图边界
黄色格子是坦克出生点, 如果放了墙会让坦克卡住

设置好表格后运行map_generator.py
把打印出来的结果复制
在setting里面仿照之前的地图赋值一个变量. e.g.
WALL_POSITIONS3 = [((250, 0), (50, 50)), ((550, 0), (50, 50)),
                    ((250, 50), (50, 50)), ((550, 50), (50, 50)), ((700, 100), (50, 50)), 
                    ((650, 150), (50, 50)), ((250, 200), (50, 50)), ((450, 200), (50, 50)),
                      ((650, 200), (50, 50)), ((450, 250), (50, 50)), ((250, 300), (50, 50)),
                        ((450, 300), (50, 50)), ((600, 400), (50, 50)), ((100, 450), (50, 50)),
                          ((350, 450), (50, 50)), ((50, 500), (50, 50)), ((300, 500), (50, 50)),
                            ((0, 550), (50, 50)), ((250, 550), (50, 50)), ((600, 550), (50, 50))]
然后把地图加入到地图池子
"""
import openpyxl

def find_cells_with_value(file_path, value=1):
    # 打开Excel文件

    workbook = openpyxl.load_workbook(file_path)
    
    # 获取所有工作表的名称

    sheet_names = workbook.sheetnames

    
    # 用于存储包含值为1的单元格的坐标列表

    WALL_POSITIONS2 = []
    
    # 遍历每个工作表

    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        
        # 遍历工作表中的每个单元格

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == value:
                    # 将单元格的坐标添加到列表中
                    xysize = (((ord(cell.coordinate[0])-ord("A"))*50, (int(cell.coordinate[1:])-1)*50),(50, 50))
                    WALL_POSITIONS2.append(xysize)
    
    return WALL_POSITIONS2

# 示例使用

file_path = 'map_template.xlsx'  # 替换为你的Excel文件路径

cells_with_one = find_cells_with_value(file_path)

# 打印结果

print(cells_with_one)